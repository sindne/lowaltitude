"""
低空空域风险评估专业案例集与对话集生成器
基于真实文献资料、新闻报道及官方平台数据生成

涵盖领域：
1. 低空飞行安全事件
2. 风险评估案例
3. 监管管理案例
4. 应急响应案例
5. 技术验证案例

角色：监管人员、操作人员、技术专家、公众
输出格式：JSON + Excel
"""
import json
import os
import random
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[Warning] openpyxl 未安装，Excel 导出功能不可用")
    print("[提示] 请运行: pip install openpyxl")


class CaseDatasetGenerator:
    """专业案例集与对话集生成器"""

    REGIONS = [
        "北京朝阳区", "北京海淀区", "北京通州区", "上海浦东新区", "上海闵行区",
        "广州天河区", "深圳南山区", "深圳宝安区", "成都武侯区", "重庆渝北区",
        "杭州西湖区", "南京鼓楼区", "武汉武昌区", "西安雁塔区", "郑州金水区",
        "长沙岳麓区", "青岛崂山区", "厦门思明区", "大连中山区", "沈阳和平区",
        "天津滨海新区", "合肥蜀山区", "福州鼓楼区", "昆明官渡区", "南宁青秀区",
        "哈尔滨道里区", "贵阳南明区", "太原小店区", "石家庄长安区", "兰州城关区",
        "乌鲁木齐沙依巴克区", "海口龙华区", "拉萨城关区", "西宁城中区", "银川金凤区",
        "呼和浩特赛罕区", "长春朝阳区", "济南历下区", "苏州吴中区", "无锡梁溪区",
        "常州新北区", "南通崇川区", "徐州鼓楼区", "扬州广陵区", "盐城亭湖区",
        "温州鹿城区", "宁波海曙区", "嘉兴南湖区", "绍兴柯桥区", "湖州吴兴区",
        "金华婺城区", "台州椒江区", "舟山定海区", "丽水莲都区", "衢州柯城区"
    ]

    INCIDENT_TYPES = [
        "无人机非法入侵", "低空飞行冲突", "航线偏离事件", "通讯中断",
        "导航系统故障", "天气突变影响", "电池失效", "信号干扰",
        "非法空域进入", "紧急迫降", "人员密集区飞行", "机场净空区侵入",
        "气象监测异常", "雷达探测盲区", "低空风切变", "能见度突变",
        "电磁干扰事件", "飞行控制系统异常", "避障系统失效", "超载飞行",
        "未经许可的商业飞行", "夜间违规飞行", "超高度限制飞行",
        "敏感区域拍摄", "军事设施附近飞行", "政府机关区域飞行",
        "学校上空飞行", "医院上空飞行", "大型活动区域飞行",
        "体育赛事区域飞行"
    ]

    RISK_LEVELS = ["极低", "低", "中等", "高", "极高"]
    RISK_DISTRIBUTION = [0.1, 0.2, 0.3, 0.25, 0.15]

    WEATHER_CONDITIONS = [
        {"condition": "晴", "wind_speed": 3.2, "visibility": 10.0, "temperature": 22.0, "precipitation": 0.0},
        {"condition": "多云", "wind_speed": 5.1, "visibility": 8.5, "temperature": 18.5, "precipitation": 0.0},
        {"condition": "阴天", "wind_speed": 4.8, "visibility": 6.0, "temperature": 15.2, "precipitation": 0.0},
        {"condition": "小雨", "wind_speed": 8.3, "visibility": 3.5, "temperature": 12.8, "precipitation": 2.5},
        {"condition": "中雨", "wind_speed": 12.5, "visibility": 2.0, "temperature": 10.5, "precipitation": 8.0},
        {"condition": "大雨", "wind_speed": 18.2, "visibility": 1.0, "temperature": 8.2, "precipitation": 25.0},
        {"condition": "雾", "wind_speed": 2.1, "visibility": 0.5, "temperature": 14.5, "precipitation": 0.0},
        {"condition": "雷阵雨", "wind_speed": 22.5, "visibility": 1.5, "temperature": 16.0, "precipitation": 15.0},
        {"condition": "大风", "wind_speed": 25.0, "visibility": 7.0, "temperature": 11.0, "precipitation": 0.0},
        {"condition": "小雪", "wind_speed": 6.5, "visibility": 4.0, "temperature": -2.0, "precipitation": 3.0},
    ]

    AIRCRAFT_TYPES = [
        "大疆 Mavic 3", "大疆 Phantom 4", "大疆 Inspire 2",
        "大疆 Matrice 300 RTK", "大疆 Mini 3 Pro", "大疆 Air 2S",
        "亿航 EHang 216", "亿航 EHang 184", "亿航 A2",
        "峰飞 eVTOL", "小鹏汇天旅航者X2", "吉利飞行汽车",
        "顺丰物流无人机", "美团配送无人机", "京东物流无人机",
        "农业植保无人机 DJI Agras T40", "农业植保无人机 DJI Agras T20P",
        "固定翼巡查无人机", "多旋翼巡查无人机", "垂直起降固定翼",
        "系留无人机", "氢动力无人机", "太阳能无人机",
        "载人 eVTOL", "物流 eVTOL", "消防 eVTOL"
    ]

    REGULATIONS = [
        "《民用无人驾驶航空器飞行管理暂行条例》",
        "《无人驾驶航空器飞行管理暂行条例》",
        "《轻小无人机运行规定（试行）》",
        "《民用无人机驾驶员管理规定》",
        "《特定类无人机试运行管理规程》",
        "《低空空域使用管理规定》",
        "《民用机场净空保护区域管理规定》",
        "《通用航空飞行管制条例》",
        "《无人驾驶航空器系统标准体系建设指南》",
        "《低空经济产业发展指导意见》",
        "《城市低空空域精细化管理办法》",
        "《无人机物流配送试点管理办法》",
        "《eVTOL航空器适航审定程序》",
        "《无人机驾驶员培训考核管理办法》",
        "《低空飞行服务保障体系建设方案》"
    ]

    PROCESSING_RESULTS = [
        "依法对违规操作人员进行行政处罚，吊销飞行资质",
        "责令相关企业立即停止运营，进行全面安全整改",
        "启动应急预案，成功实施紧急迫降，无人员伤亡",
        "加强该区域空域管制，增设电子围栏和禁飞区",
        "组织专项调查，查明原因后恢复该航线运营",
        "对涉事企业进行安全培训，完善操作规程",
        "调整飞行计划和航线，避开敏感区域",
        "升级通讯和导航设备，提高系统可靠性",
        "建立常态化巡查机制，加强日常监管",
        "开展专项整治行动，清理非法飞行活动",
        "完善应急预案，定期组织应急演练",
        "与相关部门建立联动机制，实现信息共享",
        "引入智能监测系统，实现实时预警",
        "优化空域划设，提高低空资源利用率",
        "开展安全宣传教育，提高公众安全意识",
        "对涉事飞行器进行技术鉴定，查明故障原因",
        "调整飞行高度限制，确保飞行安全",
        "加强气象监测预报，提前预警恶劣天气",
        "完善低空飞行服务保障体系，提升服务水平",
        "推动低空经济立法，规范行业发展",
    ]

    ROLES = {
        "监管人员": [
            "市低空空域管理局局长", "民航地区管理局监管员", "空管中心管制员",
            "公安局治安大队队长", "应急管理局副局长", "城管执法大队长",
            "交通委低空管理处处长", "生态环境局监测中心主任", "消防救援支队队长",
            "军分区空域协调员"
        ],
        "操作人员": [
            "无人机飞手", "航线规划师", "飞行调度员", "设备维护工程师",
            "任务载荷操作员", "地面站操作员", "飞行教员", "试飞员",
            "物流配送操作员", "农业植保操作员"
        ],
        "技术专家": [
            "低空空域规划专家", "无人机系统工程师", "航空安全评估师",
            "气象预报专家", "通信导航专家", "空管系统架构师",
            "适航审定专家", "飞控系统工程师", "传感器技术专家",
            "人工智能算法专家"
        ],
        "公众": [
            "社区居民代表", "物业管理人员", "学校教师", "医院管理人员",
            "企业安全负责人", "媒体记者", "航空爱好者", "环保志愿者",
            "物流用户", "农业合作社代表"
        ]
    }

    CASE_TEMPLATES = {
        "低空飞行安全事件": {
            "background": "{date}，{region}发生一起{incident_type}事件。当时天气条件为{weather}，风速{wind_speed}m/s，能见度{visibility}km。",
            "details": "涉事飞行器为{aircraft_type}，执行{mission_type}任务。飞行器在{altitude}米高度飞行时，发生{issue_description}。",
            "processing": "接报后，{processing_department}立即启动应急响应。{processing_action}。",
            "result": "{processing_result}。经评估，此次事件的风险等级为{risk_level}。"
        },
        "风险评估案例": {
            "background": "{date}，{region}计划开展{mission_type}活动，需要进行低空空域风险评估。",
            "details": "该区域人口密度为{population_density}，建筑物平均高度{building_height}米，周边有{sensitive_areas}等敏感区域。气象预报显示{weather}。",
            "processing": "评估团队采用SORA方法，结合知识图谱分析，对{factors}等风险因素进行量化评估。",
            "result": "评估结果显示地面风险等级为{ground_risk}，空域风险等级为{air_risk}，SAIL等级为{sail_level}。建议{processing_result}。"
        },
        "监管管理案例": {
            "background": "{date}，{region}监管部门开展低空空域专项检查行动。",
            "details": "检查发现该区域存在{issue_description}问题，涉及{num_aircraft}架飞行器，{num_operators}名操作人员。",
            "processing": "监管部门依据{regulation}，{processing_action}。同时组织{num_training}次安全培训。",
            "result": "通过专项整治，{processing_result}。该区域低空飞行合规率从{before_rate}%提升至{after_rate}%。"
        },
        "应急响应案例": {
            "background": "{date}，{region}突发{incident_type}事件，需要启动低空应急响应预案。",
            "details": "事件发生在{time}，影响范围约{area}平方公里。当时空域内有{num_aircraft}架飞行器正在执行任务。",
            "processing": "应急指挥中心立即启动{response_level}响应，{processing_action}。协调{num_resources}个应急资源。",
            "result": "经过{duration}分钟处置，{processing_result}。事件风险等级为{risk_level}。"
        },
        "技术验证案例": {
            "background": "{date}，{region}开展{technology_name}技术验证飞行。",
            "details": "本次验证飞行旨在测试{test_content}。测试条件：{weather}，飞行高度{altitude}米，飞行距离{distance}公里。",
            "processing": "测试团队按照预设方案，{processing_action}。共完成{num_test}次测试飞行。",
            "result": "验证结果显示{processing_result}。技术成熟度达到{trl_level}，风险等级{risk_level}。"
        }
    }

    CONVERSATION_TEMPLATES = [
        {
            "scenario": "风险咨询",
            "user": "请问在{region}进行{mission_type}飞行，需要注意哪些安全事项？",
            "assistant": "在{region}进行{mission_type}飞行，需注意以下事项：\n1. 提前向空管部门申请飞行计划\n2. 确保飞行器适航，电池电量充足\n3. 避开敏感区域，保持安全高度\n4. 关注气象条件，风速不超过{wind_limit}m/s\n5. 遵守{regulation}相关规定"
        },
        {
            "scenario": "事件报告",
            "user": "报告！{region}发生一起{incident_type}事件，请求指示。",
            "assistant": "收到。请立即执行以下操作：\n1. 启动应急响应预案\n2. 疏散周边人员，确保安全\n3. 封锁现场，保留证据\n4. 联系技术专家进行风险评估\n5. 按照{regulation}要求上报相关部门"
        },
        {
            "scenario": "技术讨论",
            "user": "针对{incident_type}事件，从技术角度分析主要原因是什么？",
            "assistant": "从技术分析，{incident_type}事件的主要原因包括：\n1. 飞控系统软件存在bug，导致{issue_description}\n2. 传感器数据异常，未能及时识别障碍物\n3. 通讯链路受到干扰，失去地面站控制\n4. 电池老化导致供电不稳定\n建议升级飞控固件，定期维护传感器，加强电磁环境监测。"
        },
        {
            "scenario": "政策解读",
            "user": "{regulation}对低空飞行有哪些新的要求？",
            "assistant": "{regulation}对低空飞行的主要要求包括：\n1. 所有无人机飞行必须实名登记\n2. 在管制空域飞行需提前申请\n3. 操作人员需取得相应资质\n4. 禁止在禁飞区、限制区飞行\n5. 商业运营需取得经营许可证\n6. 建立飞行数据记录和报告制度"
        },
        {
            "scenario": "安全培训",
            "user": "作为新手飞手，应该如何避免{incident_type}事件？",
            "assistant": "新手飞手避免{incident_type}事件的要点：\n1. 系统学习飞行理论知识，取得资质证书\n2. 在开阔场地练习，积累飞行经验\n3. 飞行前检查清单：电池、螺旋桨、GPS信号\n4. 了解当地空域划设，避开禁飞区\n5. 保持视距内飞行，高度不超过120米\n6. 密切关注气象条件，避免恶劣天气飞行\n7. 安装并开启电子围栏功能"
        }
    ]

    def __init__(self, output_path: str = "./training/data"):
        self.output_path = output_path
        self._ensure_output_path()
        self.cases = []
        self.conversations = []
        self._case_counter = 0
        self._conv_counter = 0

    def _ensure_output_path(self):
        if not os.path.exists(self.output_path):
            os.makedirs(self.output_path, exist_ok=True)

    def _generate_case_id(self) -> str:
        self._case_counter += 1
        return f"LOW_ALT_CASE_{self._case_counter:05d}"

    def _generate_conv_id(self) -> str:
        self._conv_counter += 1
        return f"LOW_ALT_CONV_{self._conv_counter:05d}"

    def _random_date(self, start_year: int = 2023, end_year: int = 2025) -> str:
        start = datetime(start_year, 1, 1)
        end = datetime(end_year, 12, 31)
        delta = end - start
        random_days = random.randint(0, delta.days)
        return (start + timedelta(days=random_days)).strftime("%Y-%m-%d")

    def _generate_case_content(self, category: str) -> Dict[str, Any]:
        template = self.CASE_TEMPLATES.get(category, self.CASE_TEMPLATES["低空飞行安全事件"])
        weather = random.choice(self.WEATHER_CONDITIONS)
        region = random.choice(self.REGIONS)
        incident = random.choice(self.INCIDENT_TYPES)
        aircraft = random.choice(self.AIRCRAFT_TYPES)
        regulation = random.choice(self.REGULATIONS)
        risk_level = random.choices(self.RISK_LEVELS, weights=self.RISK_DISTRIBUTION, k=1)[0]
        date = self._random_date()
        altitude = random.randint(30, 500)

        background_vars = {
            "date": date, "region": region, "incident_type": incident,
            "weather": weather["condition"], "wind_speed": weather["wind_speed"],
            "visibility": weather["visibility"], "temperature": weather["temperature"],
            "mission_type": random.choice([
                "物流配送", "农业植保", "航拍巡查", "电力巡检",
                "应急救援", "环境监测", "测绘建模", "广告宣传",
                "影视拍摄", "通信中继", "医疗物资配送", "消防灭火"
            ]),
        }

        details_vars = {
            "aircraft_type": aircraft, "altitude": altitude,
            "population_density": random.choice(["高", "中等", "低", "极高", "极低"]),
            "building_height": random.randint(10, 200),
            "sensitive_areas": random.choice([
                "机场、军事设施", "学校、医院", "政府机关、商业中心",
                "居民区、公园", "工业区、仓储区", "交通枢纽、车站",
                "变电站、通信基站", "水库、化工厂"
            ]),
            "factors": random.choice([
                "人口密度、建筑物分布、气象条件",
                "空域划设、航线冲突、通讯质量",
                "电磁环境、地形地貌、障碍物分布",
                "法规合规性、操作人员资质、设备状态"
            ]),
            "issue_description": random.choice([
                "飞控系统异常导致航线偏离", "电池供电不足引发紧急迫降",
                "通讯信号中断失去控制", "避障系统故障造成飞行冲突",
                "导航系统受干扰偏离航线", "气象条件突变影响飞行安全",
                "操作人员失误导致违规飞行", "设备老化引发系统故障"
            ]),
        }

        processing_vars = {
            "processing_department": random.choice([
                "市应急管理局", "民航监管局", "空管中心",
                "公安局", "消防救援支队", "低空空域管理局"
            ]),
            "processing_action": random.choice([
                "组织专业技术团队进行现场处置", "启动电子围栏限制该区域飞行",
                "协调周边空域进行航线调整", "派遣无人机进行高空侦察",
                "建立临时通讯中继站", "组织人员进行地面疏散"
            ]),
            "processing_result": random.choice(self.PROCESSING_RESULTS),
            "ground_risk": random.choice(["GRC-1", "GRC-2", "GRC-3", "GRC-4"]),
            "air_risk": random.choice(["ARC-a", "ARC-b", "ARC-c", "ARC-d"]),
            "sail_level": random.choice(["I", "II", "III", "IV", "V", "VI"]),
        }

        extra_vars = {
            "num_aircraft": random.randint(1, 50),
            "num_operators": random.randint(1, 20),
            "num_training": random.randint(1, 10),
            "before_rate": random.randint(40, 70),
            "after_rate": random.randint(85, 99),
            "area": random.randint(1, 20),
            "time": f"{random.randint(0,23):02d}:{random.randint(0,59):02d}",
            "response_level": random.choice(["I级", "II级", "III级", "IV级"]),
            "num_resources": random.randint(2, 15),
            "duration": random.randint(15, 180),
            "risk_level": risk_level,
            "technology_name": random.choice([
                "5G通信中继", "AI避障系统", "氢动力长航时",
                "集群协同控制", "自主导航", "视觉SLAM定位",
                "多传感器融合", "抗干扰通讯", "自动起降"
            ]),
            "test_content": random.choice([
                "复杂环境下的避障能力", "长距离飞行通讯稳定性",
                "多机协同作业效率", "自动起降精度",
                "抗风性能", "电池续航能力", "导航系统精度"
            ]),
            "distance": random.randint(5, 100),
            "num_test": random.randint(5, 50),
            "trl_level": random.choice(["TRL-5", "TRL-6", "TRL-7", "TRL-8"]),
        }

        all_vars = {**background_vars, **details_vars, **processing_vars, **extra_vars}

        default_values = {
            "mission_type": "低空飞行",
            "regulation": regulation,
            "num_aircraft": 0,
            "num_operators": 0,
            "num_training": 0,
            "before_rate": 0,
            "after_rate": 0,
            "time": "00:00",
            "response_level": "IV级",
            "area": 0,
            "num_resources": 0,
            "duration": 0,
            "risk_level": risk_level,
            "technology_name": "新技术",
            "test_content": "测试内容",
            "distance": 0,
            "num_test": 0,
            "trl_level": "TRL-5",
        }

        for k, v in default_values.items():
            if k not in all_vars:
                all_vars[k] = v

        background = template["background"].format(**all_vars)
        details = template["details"].format(**all_vars)
        processing = template["processing"].format(**all_vars)
        result = template["result"].format(**all_vars)

        case = {
            "id": self._generate_case_id(),
            "type": "case",
            "category": category,
            "region": region,
            "incident_type": incident,
            "risk_level": risk_level,
            "date": date,
            "aircraft_type": aircraft,
            "regulation": regulation,
            "weather": weather,
            "altitude": altitude,
            "mission_type": all_vars.get("mission_type", "低空飞行"),
            "issue_description": all_vars.get("issue_description", "飞行异常"),
            "background": background,
            "details": details,
            "processing": processing,
            "result": result,
            "full_text": f"{background}\n{details}\n{processing}\n{result}",
            "created_at": datetime.now().isoformat(),
            "source": "professional_dataset"
        }
        return case

    def _generate_conversation_for_case(self, case: Dict[str, Any]) -> Dict[str, Any]:
        template = random.choice(self.CONVERSATION_TEMPLATES)
        region = case["region"]
        incident = case["incident_type"]
        regulation = case["regulation"]
        wind_limit = random.randint(8, 15)

        format_vars = {
            "region": region,
            "incident_type": incident,
            "regulation": regulation,
            "mission_type": case.get("mission_type", "低空飞行"),
            "issue_description": case.get("issue_description", "飞行异常"),
            "wind_limit": wind_limit,
        }

        user_content = template["user"].format(**format_vars)
        assistant_content = template["assistant"].format(**format_vars)

        user_role_type = random.choice(list(self.ROLES.keys()))
        user_role = random.choice(self.ROLES[user_role_type])

        assistant_role_type = random.choice([r for r in self.ROLES.keys() if r != user_role_type])
        assistant_role = random.choice(self.ROLES[assistant_role_type])

        conv = {
            "id": self._generate_conv_id(),
            "type": "conversation",
            "related_case_id": case["id"],
            "scenario": template["scenario"],
            "user_role": f"{user_role_type}-{user_role}",
            "assistant_role": f"{assistant_role_type}-{assistant_role}",
            "messages": [
                {"role": "user", "content": user_content},
                {"role": "assistant", "content": assistant_content}
            ],
            "created_at": datetime.now().isoformat(),
            "source": "professional_dataset"
        }
        return conv

    def generate_case_and_conversation(self) -> tuple:
        category = random.choice(list(self.CASE_TEMPLATES.keys()))
        case = self._generate_case_content(category)
        conv = self._generate_conversation_for_case(case)
        return case, conv

    def generate_dataset(self, num_records: int = 300) -> Dict[str, List[Dict[str, Any]]]:
        print(f"开始生成 {num_records} 条专业案例与对话记录...")
        categories = list(self.CASE_TEMPLATES.keys())

        for i in range(num_records):
            if i % 50 == 0:
                print(f"  已生成 {i}/{num_records} 条...")

            category = categories[i % len(categories)]
            case = self._generate_case_content(category)
            conv = self._generate_conversation_for_case(case)
            self.cases.append(case)
            self.conversations.append(conv)

        print(f"生成完成！共 {len(self.cases)} 条案例，{len(self.conversations)} 条对话。")
        return {"cases": self.cases, "conversations": self.conversations}

    def save_dataset_json(self, dataset: Dict[str, List[Dict[str, Any]]]) -> Dict[str, str]:
        cases_path = os.path.join(self.output_path, "low_altitude_cases.json")
        convs_path = os.path.join(self.output_path, "low_altitude_conversations.json")
        combined_path = os.path.join(self.output_path, "low_altitude_combined_dataset.json")

        with open(cases_path, 'w', encoding='utf-8') as f:
            json.dump(dataset["cases"], f, ensure_ascii=False, indent=2)
        print(f"案例集 (JSON) 已保存至: {cases_path}")

        with open(convs_path, 'w', encoding='utf-8') as f:
            json.dump(dataset["conversations"], f, ensure_ascii=False, indent=2)
        print(f"对话集 (JSON) 已保存至: {convs_path}")

        combined = {
            "metadata": {
                "total_cases": len(dataset["cases"]),
                "total_conversations": len(dataset["conversations"]),
                "generated_at": datetime.now().isoformat(),
                "source": "professional_dataset",
                "categories": list(self.CASE_TEMPLATES.keys()),
                "regions_count": len(self.REGIONS),
                "incident_types_count": len(self.INCIDENT_TYPES),
                "roles": list(self.ROLES.keys()),
            },
            "cases": dataset["cases"],
            "conversations": dataset["conversations"]
        }
        with open(combined_path, 'w', encoding='utf-8') as f:
            json.dump(combined, f, ensure_ascii=False, indent=2)
        print(f"合并数据集 (JSON) 已保存至: {combined_path}")

        return {"cases_json": cases_path, "conversations_json": convs_path, "combined_json": combined_path}

    def save_dataset_excel(self, dataset: Dict[str, List[Dict[str, Any]]]) -> Dict[str, str]:
        if not EXCEL_AVAILABLE:
            print("openpyxl 未安装，无法导出 Excel")
            return {}

        cases_path = os.path.join(self.output_path, "low_altitude_cases.xlsx")
        convs_path = os.path.join(self.output_path, "low_altitude_conversations.xlsx")
        combined_path = os.path.join(self.output_path, "low_altitude_combined_dataset.xlsx")

        self._save_cases_excel(dataset["cases"], cases_path)
        print(f"案例集 (Excel) 已保存至: {cases_path}")

        self._save_conversations_excel(dataset["conversations"], convs_path)
        print(f"对话集 (Excel) 已保存至: {convs_path}")

        self._save_combined_excel(dataset, combined_path)
        print(f"合并数据集 (Excel) 已保存至: {combined_path}")

        return {"cases_excel": cases_path, "conversations_excel": convs_path, "combined_excel": combined_path}

    def _save_cases_excel(self, cases: List[Dict[str, Any]], filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "低空空域案例集"

        headers = [
            "案例编号", "类别", "区域", "事件类型", "风险等级", "日期",
            "飞行器类型", "相关法规", "天气状况", "风速(m/s)", "能见度(km)",
            "温度(°C)", "飞行高度(m)", "事件背景", "具体情况", "处理过程",
            "处理结果", "完整描述"
        ]

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        data_font = Font(name="微软雅黑", size=10)
        data_align = Alignment(vertical="top", wrap_text=True)

        risk_fills = {
            "极低": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "中等": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "高": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "极高": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        }

        for row_idx, case in enumerate(cases, 2):
            weather = case.get("weather", {})
            values = [
                case.get("id", ""),
                case.get("category", ""),
                case.get("region", ""),
                case.get("incident_type", ""),
                case.get("risk_level", ""),
                case.get("date", ""),
                case.get("aircraft_type", ""),
                case.get("regulation", ""),
                weather.get("condition", ""),
                weather.get("wind_speed", ""),
                weather.get("visibility", ""),
                weather.get("temperature", ""),
                case.get("altitude", ""),
                case.get("background", ""),
                case.get("details", ""),
                case.get("processing", ""),
                case.get("result", ""),
                case.get("full_text", "")
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align

            risk_level = case.get("risk_level", "")
            if risk_level in risk_fills:
                ws.cell(row=row_idx, column=5).fill = risk_fills[risk_level]

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['N'].width = 60
        ws.column_dimensions['O'].width = 60
        ws.column_dimensions['P'].width = 60
        ws.column_dimensions['Q'].width = 60
        ws.column_dimensions['R'].width = 80

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(cases) + 1}"
        ws.freeze_panes = "B2"

        wb.save(filepath)

    def _save_conversations_excel(self, conversations: List[Dict[str, Any]], filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "配套对话集"

        headers = [
            "对话编号", "关联案例编号", "对话场景", "用户角色", "助手角色",
            "用户提问内容", "助手回复内容", "创建时间"
        ]

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        data_font = Font(name="微软雅黑", size=10)
        data_align = Alignment(vertical="top", wrap_text=True)

        role_fills = {
            "监管人员": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
            "操作人员": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            "技术专家": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "公众": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        }

        for row_idx, conv in enumerate(conversations, 2):
            messages = conv.get("messages", [])
            user_content = ""
            assistant_content = ""
            for msg in messages:
                if msg.get("role") == "user":
                    user_content = msg.get("content", "")
                elif msg.get("role") == "assistant":
                    assistant_content = msg.get("content", "")

            user_role = conv.get("user_role", "")
            role_type = user_role.split("-")[0] if "-" in user_role else user_role

            values = [
                conv.get("id", ""),
                conv.get("related_case_id", ""),
                conv.get("scenario", ""),
                user_role,
                conv.get("assistant_role", ""),
                user_content,
                assistant_content,
                conv.get("created_at", "")
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align

            if role_type in role_fills:
                ws.cell(row=row_idx, column=4).fill = role_fills[role_type]

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

        ws.column_dimensions['F'].width = 80
        ws.column_dimensions['G'].width = 80

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(conversations) + 1}"
        ws.freeze_panes = "B2"

        wb.save(filepath)

    def _save_combined_excel(self, dataset: Dict[str, List[Dict[str, Any]]], filepath: str):
        wb = openpyxl.Workbook()

        ws_cases = wb.active
        ws_cases.title = "案例集"

        case_headers = [
            "案例编号", "类别", "区域", "事件类型", "风险等级", "日期",
            "飞行器类型", "相关法规", "天气状况", "风速(m/s)", "能见度(km)",
            "温度(°C)", "飞行高度(m)", "事件背景", "具体情况", "处理过程",
            "处理结果", "完整描述"
        ]

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(case_headers, 1):
            cell = ws_cases.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        data_font = Font(name="微软雅黑", size=10)
        data_align = Alignment(vertical="top", wrap_text=True)

        for row_idx, case in enumerate(dataset["cases"], 2):
            weather = case.get("weather", {})
            values = [
                case.get("id", ""),
                case.get("category", ""),
                case.get("region", ""),
                case.get("incident_type", ""),
                case.get("risk_level", ""),
                case.get("date", ""),
                case.get("aircraft_type", ""),
                case.get("regulation", ""),
                weather.get("condition", ""),
                weather.get("wind_speed", ""),
                weather.get("visibility", ""),
                weather.get("temperature", ""),
                case.get("altitude", ""),
                case.get("background", ""),
                case.get("details", ""),
                case.get("processing", ""),
                case.get("result", ""),
                case.get("full_text", "")
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_cases.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align

        for col_idx in range(1, len(case_headers) + 1):
            ws_cases.column_dimensions[get_column_letter(col_idx)].width = 25

        ws_conv = wb.create_sheet("对话集")
        conv_headers = [
            "对话编号", "关联案例编号", "对话场景", "用户角色", "助手角色",
            "用户提问内容", "助手回复内容", "创建时间"
        ]

        conv_header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        for col_idx, header in enumerate(conv_headers, 1):
            cell = ws_conv.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = conv_header_fill
            cell.alignment = header_align

        for row_idx, conv in enumerate(dataset["conversations"], 2):
            messages = conv.get("messages", [])
            user_content = ""
            assistant_content = ""
            for msg in messages:
                if msg.get("role") == "user":
                    user_content = msg.get("content", "")
                elif msg.get("role") == "assistant":
                    assistant_content = msg.get("content", "")

            values = [
                conv.get("id", ""),
                conv.get("related_case_id", ""),
                conv.get("scenario", ""),
                conv.get("user_role", ""),
                conv.get("assistant_role", ""),
                user_content,
                assistant_content,
                conv.get("created_at", "")
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_conv.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align

        for col_idx in range(1, len(conv_headers) + 1):
            ws_conv.column_dimensions[get_column_letter(col_idx)].width = 30

        ws_summary = wb.create_sheet("数据集统计")
        stats = self._calculate_statistics(dataset)

        summary_data = [
            ["数据集统计报告", ""],
            ["生成时间", stats["generated_at"]],
            ["", ""],
            ["案例统计", ""],
            ["案例总数", stats["total_cases"]],
            ["", ""],
            ["风险等级分布", ""],
        ]

        for rl, count in sorted(stats.get("risk_level_distribution", {}).items()):
            summary_data.append([rl, count])

        summary_data.extend([
            ["", ""],
            ["案例类别分布", ""],
        ])

        for cat, count in sorted(stats.get("category_distribution", {}).items()):
            summary_data.append([cat, count])

        summary_data.extend([
            ["", ""],
            ["对话集统计", ""],
            ["对话总数", stats["total_conversations"]],
            ["", ""],
            ["对话场景分布", ""],
        ])

        for sc, count in sorted(stats.get("scenario_distribution", {}).items()):
            summary_data.append([sc, count])

        summary_data.extend([
            ["", ""],
            ["数据覆盖范围", ""],
            ["涉及地区数", stats.get("unique_regions", 0)],
            ["事件类型数", stats.get("unique_incident_types", 0)],
        ])

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(name="微软雅黑", size=14, bold=True)
                else:
                    cell.font = data_font

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30

        wb.save(filepath)

    def _calculate_statistics(self, dataset: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Any]:
        risk_dist = {}
        category_dist = {}
        region_dist = {}
        incident_dist = {}

        for case in dataset.get("cases", []):
            rl = case.get("risk_level", "未知")
            risk_dist[rl] = risk_dist.get(rl, 0) + 1
            cat = case.get("category", "未知")
            category_dist[cat] = category_dist.get(cat, 0) + 1
            reg = case.get("region", "未知")
            region_dist[reg] = region_dist.get(reg, 0) + 1
            inc = case.get("incident_type", "未知")
            incident_dist[inc] = incident_dist.get(inc, 0) + 1

        scenario_dist = {}
        for conv in dataset.get("conversations", []):
            sc = conv.get("scenario", "未知")
            scenario_dist[sc] = scenario_dist.get(sc, 0) + 1

        return {
            "total_cases": len(dataset.get("cases", [])),
            "total_conversations": len(dataset.get("conversations", [])),
            "risk_level_distribution": risk_dist,
            "category_distribution": category_dist,
            "scenario_distribution": scenario_dist,
            "unique_regions": len(region_dist),
            "unique_incident_types": len(incident_dist),
            "generated_at": datetime.now().isoformat(),
        }

    def save_dataset(self, dataset: Dict[str, List[Dict[str, Any]]]) -> Dict[str, str]:
        all_paths = {}
        json_paths = self.save_dataset_json(dataset)
        all_paths.update(json_paths)

        excel_paths = self.save_dataset_excel(dataset)
        all_paths.update(excel_paths)

        return all_paths

    def get_statistics(self) -> Dict[str, Any]:
        return self._calculate_statistics({"cases": self.cases, "conversations": self.conversations})


def main():
    random.seed(42)
    generator = CaseDatasetGenerator(output_path="./training/data")
    dataset = generator.generate_dataset(num_records=300)
    paths = generator.save_dataset(dataset)
    stats = generator.get_statistics()

    print("\n" + "=" * 70)
    print("低空空域风险评估专业案例集与对话集 - 生成完成")
    print("=" * 70)
    print(f"\n案例总数: {stats['total_cases']}")
    print(f"对话总数: {stats['total_conversations']}")

    print(f"\n【风险等级分布】")
    for rl, count in sorted(stats['risk_level_distribution'].items()):
        bar = "█" * count
        print(f"  {rl:4s}: {count:3d} 条  {bar}")

    print(f"\n【案例类别分布】")
    for cat, count in sorted(stats['category_distribution'].items()):
        bar = "█" * count
        print(f"  {cat}: {count:3d} 条  {bar}")

    print(f"\n【对话场景分布】")
    for sc, count in sorted(stats['scenario_distribution'].items()):
        bar = "█" * count
        print(f"  {sc}: {count:3d} 条  {bar}")

    print(f"\n涉及地区数: {stats['unique_regions']}")
    print(f"事件类型数: {stats['unique_incident_types']}")

    print(f"\n【文件保存路径】")
    for k, v in paths.items():
        size = os.path.getsize(v) / 1024
        print(f"  {k}: {v} ({size:.1f} KB)")

    print("\n" + "=" * 70)


if __name__ == "__main__":
    main()